#!/usr/bin/env python3

# For Network Rates
# This script will convert each tab in the Network Excel file provided by the Product Owner into their own sql file
# NOTE: This will also add the Header/Footers needed for the Oracle scripts.

from openpyxl import load_workbook
from datetime import date

filename = "./testfile/createSqlFile.xlsx"
today = date.today()

def writeHeader(tablename, outputFile):
    print("/*", file=outputFile)
    print("    Script: test.sql".format(tablename), file=outputFile)
    print("    Description: {} Source file tablename from {}".format(tablename, today), file=outputFile)
    print("    Output: ..\..\..\..\log\\{}.data.out".format(tablename), file=outputFile)
    print("*/", file=outputFile)
    print("", file=outputFile)
    print("set pagesize 1000 linesize 300 trimspool on;", file=outputFile)
    print("set echo off;", file=outputFile)
    print("set feedback off;", file=outputFile)
    print("SET DEFINE OFF;", file=outputFile)
    print("", file=outputFile)
    print("spool ..\..\..\..\log\\{}.data.out".format(tablename), file=outputFile)
    print("PROMPT Load the {} Source file tablename from {}".format(tablename, today), file=outputFile)
    print("", file=outputFile)
    print("truncate table {};".format(tablename.upper()), file=outputFile)

    print("SET DEFINE OFF", file=outputFile)
    print("", file=outputFile)

def writeFooter(outputFile):
    print("COMMIT;", file=outputFile)
    print("spool off;", file=outputFile)
    print("exit;", file=outputFile)

def writeValueToFile(cellValue, outputFile):
    if cellValue is None:
        cellValue = ""
    elif isinstance(cellValue, str):
        cellValue = cellValue.replace("'","''")
    print("'{}', ".format(cellValue), end='', file=outputFile)

def loopThrough(ws, row, lastColumn, outputfile):
    for i in range(lastColumn):
        writeValueToFile(ws.cell(row, i+1).value, outputfile)

def createSheet1File(ws):
    sheet1InsertStatement = "INSERT INTO sheet1 (value1, value2, value3, date1, date2) "
    sheet1SqlFile = open("sheet1.data.sql", "w")

    writeHeader("sheet1", sheet1SqlFile)
    for row in range(2, ws.max_row+1):
        if ws.cell(row,1).value is not None:
            print(sheet1InsertStatement, file=sheet1SqlFile)
            print("VALUES (", end='', file=sheet1SqlFile)
            loopThrough(ws, row, 4, sheet1SqlFile)
            lastvalue = ws.cell(row, 5).value
            if lastvalue is None:
                lastvalue = ""
            print("'{}');".format(lastvalue), file=sheet1SqlFile)
            print(file=sheet1SqlFile)

    writeFooter(sheet1SqlFile)
    sheet1SqlFile.close()

def main():
    wb = load_workbook(filename)
    createSheet1File(wb['test1'])

if __name__ == "__main__":
    main()

