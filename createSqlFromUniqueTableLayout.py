#!/usr/bin/env python3


# This script will convert the Excel file into sql file

from openpyxl import load_workbook
filename = "../testfile/test2.xlsx"
createDate = "'2023-11-01'"
expireDate = ",'2024-08-31',"
yearDefined = 2024

test2InsertStatement = "INSERT [sheet1] (C1, C2, C3, C4, C5, C6, C7, C8, C9, C10, C11, C12, C13, C14, C15, C16, C17) VALUES ("
test3InsertStatement = "INSERT [sheet2] (C1, C2, C3, C4, C5, C6, C7, C8, C9, C10, C11, C12, C13, C14, C15, C16, C17) VALUES ('"

def replaceNullString(value):
    if value is None:
        return ""
    if value == "Null":
        return ""
    if value == "Nu":
        return ""
    if value == "NULL":
        return ""
    return value

def replaceEmptyWith0(value):
    if value is None:
        return 0
    return value

def isImportExportRow(row, ws):
    global importExportCode
    global importExportName
    global headerRow
    if ws.cell(row, 3).value == 'Import':
        importExportCode = 1
        importExportName = "Import"
        headerRow = row
    elif ws.cell(row, 3).value == 'Export':
        importExportCode = 2
        importExportName = "Export"
        headerRow = row

def createSheet1SqlFile(ws):
    outputFile = open("sheet1.sql", "w")

    print("USE [schema]", file=outputFile)
    print("GO", file=outputFile)
    for row in range(2, ws.max_row-1):
        if ws.cell(row,1).value is not None:
            print(test2InsertStatement, end='', file=outputFile)
            print("'{}',".format(ws.cell(row, 1).value.lstrip('0')),end='', file=outputFile)
            print("'{}',".format(ws.cell(row, 2).value.lstrip('0')),end='', file=outputFile)
            print("'{}',".format(ws.cell(row, 3).value),end='', file=outputFile)
            print("{},".format(round(ws.cell(row, 4).value,2)),end='', file=outputFile) # RATE
            print("'{}',".format(ws.cell(row,5).value.strftime("%Y-%m-%d")),end='', file=outputFile) # EFFECTIVE DATE
            print("'{}',".format(ws.cell(row,6).value.strftime("%Y-%m-%d")),end='', file=outputFile) # EXPIRATION DATE
            print("'{}',".format(ws.cell(row,7).value.strftime("%b %d %Y 12:00AM")),end='', file=outputFile) # CREATED DATE
            print("'{}',".format(ws.cell(row, 8).value),end='', file=outputFile)
            print("'{}',".format(replaceNullString(ws.cell(row, 9).value)),end='', file=outputFile)
            print("'{}',".format(replaceNullString(ws.cell(row, 10).value)),end='', file=outputFile)
            print("'{}',".format(replaceNullString(ws.cell(row, 11).value)),end='', file=outputFile)
            print("'{}',".format(replaceNullString(ws.cell(row, 12).value)),end='', file=outputFile) # INACTIVE_DATE
            print("{},".format(ws.cell(row, 13).value),end='', file=outputFile) # FISCAL_YR
            print("'{}',".format(replaceNullString(ws.cell(row, 14).value)),end='', file=outputFile)
            print("'{}',".format(replaceNullString(ws.cell(row, 15).value)),end='', file=outputFile)
            print("'{}',".format(replaceNullString(ws.cell(row, 16).value)),end='', file=outputFile)
            print("'{}',".format(replaceNullString(ws.cell(row, 17).value)),end='', file=outputFile)
            print("'{}');".format(replaceEmptyWith0(ws.cell(row, 18).value)), file=outputFile)

def createSheet2SqlFile(ws):
    outputFile = open("sheet2.sql", "w")

    print("USE [schema]", file=outputFile)
    print("GO", file=outputFile)
    print("TRUNCATE TABLE [sheet2]", file=outputFile)
    print("GO", file=outputFile)
    print(file=outputFile)

    # Iterate the loop to read the cell values
    for row in range(1, ws.max_row+1):
        if ws.cell(row,1).value is None:
            isImportExportRow(row, ws)
        elif ws.cell(row,1).value == 'Code':
            # This Else If skips the Tables Header row
            skr = "skip row"
        else:
            code = ws.cell(row,1).value
            print(file=outputFile)
            print("--", importExportName, code, file=outputFile)
            for col in ws.iter_cols(3, ws.max_column):

                # Ignore any rates in the table that do not have a value
                if not col[row-1].value is None:
                    code = ws.cell(row,1).value
                    areaCode = col[headerRow].value
                    cost = col[row-1].value
                    # these prints use ,end='' to remove extra spacing
                    print (test3InsertStatement, end='', file=outputFile)
                    print(areaCode, end='', file=outputFile)
                    print("','", end='', file=outputFile)
                    print(importExportCode, end='', file=outputFile)
                    print("','MT',CAST(", end='', file=outputFile)
                    print(cost, end='', file=outputFile)
                    print(" AS Decimal(11,2)),", end='', file=outputFile)
                    print(createDate, end='', file=outputFile)
                    print(",N'####',NULL,N'NULL','", end='', file=outputFile)
                    print(code, end='', file=outputFile)
                    print("',NULL,", end='', file=outputFile)
                    print(createDate, end='', file=outputFile)
                    print(expireDate, end='', file=outputFile)
                    print(yearDefined, end='', file=outputFile)
                    print(");", file=outputFile)


def main():
    wb = load_workbook(filename)
    linerWSName = 'Sheet1'
    if linerWSName in wb.sheetnames:
        createSheet1SqlFile(wb[linerWSName])

    else:
        print("Liner Rates cannot be crated. No worksheet named '{}' in {}".format(linerWSName, filename))

    phWSName = 'Sheet2'
    if phWSName in wb.sheetnames:
        createSheet2SqlFile(wb[phWSName])
    else:
        print("PH Rates cannot be crated. No worksheet named '{}' in {}".format(phWSName, filename))


if __name__ == "__main__":
    main()